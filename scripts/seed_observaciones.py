"""Backfill the ``observaciones`` column on books and sales from Excel.

Reads the catalog Excel (``Catálogo Agosto '26 (1).xlsx``) and the sales Excel
(``Ventas 2026.xlsx``) and writes ``Book.observaciones`` / ``Sale.observaciones``
from their "Observaciones" columns. Empty/missing values default to "Juli"
("las que no digan nada son adquiridas por Juli"); non-empty values are copied
verbatim.

Sales observaciones are resolved with combined-source precedence:
  1. If the sale's first-item book natural key (title/author/editorial,
     case-insensitive) matches the SALES Excel, use that sheet's value
     (empty -> "Juli", else verbatim).
  2. Otherwise, fall back to the book's own ``observaciones`` field (already
     backfilled from the catalog in the previous step). A non-NULL value
     (including "Juli") is used.
  3. Otherwise leave ``sale.observaciones`` NULL (counted as not_found).

Sales with multiple items derive observaciones from the FIRST sale item only
(by lowest SaleItem id).

Usage:
  py scripts/seed_observaciones.py [--catalog PATH] [--sales PATH]

Defaults to the two Downloads paths above. Re-runs are safe (idempotent: the
same value is written again). The catalog path is used for books, the sales
path for sales; either flag can point at a missing file to skip that side.

Requires the backend env vars (``DATABASE_URL`` in particular — the app config
fails fast on missing required vars). When ``DATABASE_URL`` is unset, a SQLite
URL is resolved relative to the repo's ``backend/`` directory (not the CWD).
On Railway set ``DATABASE_URL`` and run the same script as a one-off. The
script fails fast if the ``books`` table is missing the ``observaciones``
column (run ``alembic upgrade head`` first).
"""

import argparse
import asyncio
import os
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "backend"))

os.environ.setdefault("SECRET_KEY", "dev-secret-not-for-production")
os.environ.setdefault("ALLOWED_ORIGINS", "http://localhost:5173,http://localhost:8000")
os.environ.setdefault("ADMIN_USERNAME", "admin")
os.environ.setdefault("ADMIN_PASSWORD", "admin")

# Resolve the default SQLite DB relative to the backend/ directory (not CWD).
BACKEND_DIR = REPO_ROOT / "backend"
DEFAULT_DATABASE_URL = (
    f"sqlite+aiosqlite:///{(BACKEND_DIR / 'bibliotheca.db').as_posix()}"
)
os.environ.setdefault("DATABASE_URL", DEFAULT_DATABASE_URL)

from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import inspect as sa_inspect  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import selectinload  # noqa: E402

from app.db import SessionLocal, engine  # noqa: E402
from app.excel_import.normalizer import (  # noqa: E402
    normalize_sheet_name,
    observaciones_column_index,
)
from app.excel_import.parser import MAX_COLUMNS, parse_workbook  # noqa: E402
from app.models import Base, Book, Sale, SaleItem  # noqa: E402
from app.services.catalog import natural_key  # noqa: E402

DEFAULT_CATALOG = Path.home() / "Downloads" / "Catálogo Agosto '26 (1).xlsx"
DEFAULT_SALES = Path.home() / "Downloads" / "Ventas 2026.xlsx"

# Only these sheets may use the JULIO column-index-5 fallback when no /observ/i
# header is present. Any other header-less sheet resolves observaciones to
# empty ("Juli") rather than assuming column 5.
JULIO_FALLBACK_SHEETS = {"ventas julio 26"}

# The JULIO sales sheet has no real header (row 1 is ``None*6, 'Juli', 'Cande'``)
# but its data rows carry the observaciones-like value at column index 5, the
# same position as the AGOSTO "OBSERVACIONES" column.
JULIO_OBS_FALLBACK_INDEX = 5


def _clean(value) -> str | None:
    """Return a stripped string cell value, or ``None`` when empty."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_catalog_observaciones(path: Path) -> dict[tuple[str, str, str], str]:
    """Map catalog natural key -> observaciones (empty falls back to "Juli").

    Reuses the catalog parser so per-sheet layout detection (genre-driven
    CATÁLOGO COMPLETO, header sheets, headerless OPORTUNIDADES) matches the
    app's own import exactly. OPORTUNIDADES has no observaciones column, so its
    rows resolve to "Juli".
    """
    parsed = parse_workbook(path)
    result: dict[tuple[str, str, str], str] = {}
    for sheet in parsed.sheets:
        for row in sheet.rows:
            if row.error is not None or row.skip_reason is not None:
                continue
            key = natural_key(row.title, row.author, row.editorial)
            result.setdefault(key, row.observaciones or "Juli")
    return result


def read_sales_observaciones(path: Path) -> dict[tuple[str, str, str], str]:
    """Map sales natural key -> observaciones (handles the JULIO header quirk).

    Only the ``ventas julio 26`` sheet uses the column-index-5 fallback when no
    /observ/i header is present. Any other header-less sheet resolves its
    observaciones to empty (which becomes "Juli" on apply).
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[tuple[str, str, str], str] = {}
    for worksheet in workbook.worksheets:
        rows_iter = worksheet.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            continue
        header = [_clean(cell) for cell in first[:MAX_COLUMNS]]
        obs_index = observaciones_column_index(header)
        if (
            obs_index is None
            and normalize_sheet_name(worksheet.title) in JULIO_FALLBACK_SHEETS
        ):
            # JULIO layout: row 1 is (None*6, 'Juli', 'Cande') with no header;
            # the observaciones-like value sits at index 5 in the data rows.
            obs_index = JULIO_OBS_FALLBACK_INDEX
        for values in rows_iter:
            cells = [_clean(cell) for cell in values[:MAX_COLUMNS]]
            if all(cell is None for cell in cells):
                continue
            title = cells[0]
            if not title:
                continue
            author = cells[1] if len(cells) > 1 else None
            editorial = cells[2] if len(cells) > 2 else None
            obs = (
                cells[obs_index]
                if obs_index is not None and obs_index < len(cells)
                else None
            )
            key = natural_key(title, author or "", editorial or "")
            result.setdefault(key, obs or "Juli")
    return result


async def backfill_book_observaciones(
    session, catalog_map: dict[tuple[str, str, str], str]
) -> tuple[int, int]:
    """Set ``Book.observaciones`` from ``catalog_map`` (only that field)."""
    updated = 0
    skipped = 0
    books = (await session.execute(select(Book))).scalars().all()
    for book in books:
        value = catalog_map.get(
            natural_key(book.title, book.author, book.editorial)
        )
        if value is None:
            skipped += 1
            continue
        book.observaciones = value
        updated += 1
    await session.commit()
    return updated, skipped


async def backfill_sale_observaciones(
    session, sales_map: dict[tuple[str, str, str], str]
) -> tuple[int, int]:
    """Set ``Sale.observaciones`` with the combined-source precedence.

    For each sale: use the sales-Excel value when the first item's book natural
    key matches; otherwise fall back to the book's own ``observaciones``; if
    neither exists, leave the sale untouched and count it as not_found.
    """
    updated = 0
    not_found = 0
    sales = (
        await session.execute(
            select(Sale)
            .options(selectinload(Sale.items).selectinload(SaleItem.book))
            .order_by(Sale.date, Sale.id)
        )
    ).scalars().all()
    for sale in sales:
        if not sale.items:
            not_found += 1
            continue
        first_item = min(sale.items, key=lambda item: item.id)
        book = first_item.book
        if book is None:
            not_found += 1
            continue
        value = sales_map.get(natural_key(book.title, book.author, book.editorial))
        if value is None:
            value = book.observaciones
        if value is None:
            not_found += 1
            continue
        sale.observaciones = value
        updated += 1
    await session.commit()
    return updated, not_found


def _books_have_observaciones(sync_conn) -> bool:
    """True when the ``books`` table already carries the ``observaciones`` column."""
    columns = {column["name"] for column in sa_inspect(sync_conn).get_columns("books")}
    return "observaciones" in columns


async def main(catalog_path: Path, sales_path: Path) -> int:
    print(f"Database URL: {engine.url}")

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    # Fail fast before any backfill: ``create_all`` never ALTERs an existing
    # table, so a pre-migration DB still lacks the column here. Surface that
    # loudly instead of silently no-op'ing or crashing mid-way.
    async with engine.connect() as conn:
        if not await conn.run_sync(_books_have_observaciones):
            print(
                "ERROR: the books table is missing the 'observaciones' column. "
                "Run `alembic upgrade head` before seeding.",
                file=sys.stderr,
            )
            return 1

    catalog_map = (
        read_catalog_observaciones(catalog_path) if catalog_path.exists() else {}
    )
    sales_map = read_sales_observaciones(sales_path) if sales_path.exists() else {}

    async with SessionLocal() as session:
        updated_books, skipped_books = await backfill_book_observaciones(
            session, catalog_map
        )
    async with SessionLocal() as session:
        updated_sales, not_found_sales = await backfill_sale_observaciones(
            session, sales_map
        )

    print(
        f"books updated={updated_books} skipped={skipped_books} "
        f"catalog_keys={len(catalog_map)}"
    )
    print(
        f"sales updated={updated_sales} not_found={not_found_sales} "
        f"sales_keys={len(sales_map)}"
    )
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Backfill observaciones on books and sales from Excel"
    )
    parser.add_argument("--catalog", type=Path, default=DEFAULT_CATALOG)
    parser.add_argument("--sales", type=Path, default=DEFAULT_SALES)
    args = parser.parse_args()
    raise SystemExit(asyncio.run(main(args.catalog, args.sales)))
