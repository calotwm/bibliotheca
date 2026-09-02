"""Tests for the ``observaciones`` field on books and sales.

Covers:
- Book create/update carrying ``observaciones``.
- Sale create + list/detail returning ``observaciones``.
- Excel import applying ``observaciones`` (verbatim, empty -> "Juli",
  Oportunidades/no-column -> "Juli").
- The ``scripts/seed_observaciones.py`` backfill logic (catalog + sales).
"""

import importlib.util
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import select

from app.excel_import.parser import parse_workbook
from app.models import Book, Category, Sale, SaleItem, User
from app.security.jwt import create_access_token
from app.security.password import hash_password

# Load the seed script as a module (scripts/ is not a package).
_SEED_PATH = Path(__file__).resolve().parents[2] / "scripts" / "seed_observaciones.py"


def _load_seed_module():
    spec = importlib.util.spec_from_file_location("seed_observaciones", _SEED_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


seed = _load_seed_module()


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def _xlsx_bytes(sheets: dict[str, list[tuple]], path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(title=name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    return path


async def _category_id(session, name="Novela") -> int:
    category = Category(name=name)
    session.add(category)
    await session.commit()
    return category.id


def _book_payload(category_id: int, **overrides) -> dict:
    payload = {
        "title": "Rayuela",
        "author": "Julio Cortázar",
        "editorial": "Sudamericana",
        "category_id": category_id,
        "price": "12.50",
        "stock": 3,
    }
    payload.update(overrides)
    return payload


async def _seed_book(
    session,
    *,
    title="Rayuela",
    author="Julio Cortázar",
    editorial="Sudamericana",
    category="Novela",
    stock=5,
) -> Book:
    cid = await _category_id(session, category)
    book = Book(
        title=title,
        author=author,
        editorial=editorial,
        category_id=cid,
        price="10.00",
        stock=stock,
    )
    session.add(book)
    await session.commit()
    return book


async def _seed_sale(session, book: Book, *, sale_number=1) -> Sale:
    sale = Sale(sale_number=sale_number, total=Decimal("10.00"))
    sale.items.append(
        SaleItem(
            book_id=book.id,
            quantity=1,
            unit_price=Decimal("10.00"),
            subtotal=Decimal("10.00"),
        )
    )
    session.add(sale)
    await session.commit()
    return sale


async def _cashier_headers(session) -> dict:
    user = User(username="cashier", password_hash=hash_password("cashier"), role="cashier")
    session.add(user)
    await session.commit()
    token = create_access_token("cashier", "cashier")
    return {"Authorization": f"Bearer {token}"}


async def _preview(client, headers, sheets):
    import io

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(title=name)
        for row in rows:
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return await client.post(
        "/api/import/preview",
        headers=headers,
        files={"file": ("catalog.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


async def _apply(client, headers, preview):
    payload = {"token": preview.json()["token"], "filename": "catalog.xlsx", "sheets": preview.json()["sheets"]}
    return await client.post("/api/import/apply", json=payload, headers=headers)


# --------------------------------------------------------------------------
# Book create/update with observaciones
# --------------------------------------------------------------------------

async def test_book_create_returns_observaciones(auth_headers, session, client):
    cid = await _category_id(session)
    response = await client.post(
        "/api/books",
        json=_book_payload(cid, observaciones="Juli y Cande"),
        headers=auth_headers,
    )
    assert response.status_code == 201
    assert response.json()["observaciones"] == "Juli y Cande"


async def test_book_create_defaults_observaciones_none(auth_headers, session, client):
    cid = await _category_id(session)
    response = await client.post(
        "/api/books", json=_book_payload(cid), headers=auth_headers
    )
    assert response.status_code == 201
    assert response.json()["observaciones"] is None


async def test_book_update_sets_observaciones(auth_headers, session, client):
    cid = await _category_id(session)
    created = await client.post(
        "/api/books", json=_book_payload(cid), headers=auth_headers
    )
    book_id = created.json()["id"]
    response = await client.put(
        f"/api/books/{book_id}",
        json={"observaciones": "Consignación Juli y Cande"},
        headers=auth_headers,
    )
    assert response.status_code == 200
    assert response.json()["observaciones"] == "Consignación Juli y Cande"

    book = (await session.execute(select(Book).where(Book.id == book_id))).scalar_one()
    assert book.observaciones == "Consignación Juli y Cande"


# --------------------------------------------------------------------------
# Sale create + list/detail with observaciones
# --------------------------------------------------------------------------

async def test_sale_create_returns_observaciones_in_detail(auth_headers, session, client):
    book = await _seed_book(session, stock=5)
    response = await client.post(
        "/api/sales",
        json={
            "items": [{"book_id": book.id, "quantity": 1}],
            "observaciones": "Juli y Cande",
        },
        headers=auth_headers,
    )
    assert response.status_code == 201
    assert response.json()["observaciones"] == "Juli y Cande"

    sale_id = response.json()["id"]
    detail = await client.get(f"/api/sales/{sale_id}", headers=auth_headers)
    assert detail.json()["observaciones"] == "Juli y Cande"


async def test_sale_list_returns_observaciones(auth_headers, session, client):
    book = await _seed_book(session, stock=5)
    await client.post(
        "/api/sales",
        json={
            "items": [{"book_id": book.id, "quantity": 1}],
            "observaciones": "Cande",
        },
        headers=auth_headers,
    )
    page = await client.get("/api/sales", headers=auth_headers)
    assert page.status_code == 200
    assert page.json()[0]["observaciones"] == "Cande"


# --------------------------------------------------------------------------
# Parser captures observaciones
# --------------------------------------------------------------------------

def test_parser_captures_observaciones_header_sheet(tmp_path):
    sheets = {
        "NOVELAS": [
            ("TÍTULO", "AUTOR", "EDITORIAL", "PRECIO", "STOCK", "OBSERVACIONES"),
            ("Rayuela", "Cortázar, Julio", "Sudamericana", 29500, 3, "Juli y Cande"),
            ("1984", "Orwell, George", "La Pollera", 24000, 1, None),
        ]
    }
    path = _xlsx_bytes(sheets, tmp_path / "catalog.xlsx")
    parsed = parse_workbook(path)
    first, second = parsed.sheets[0].rows
    assert first.observaciones == "Juli y Cande"
    assert second.observaciones is None


def test_parser_captures_observaciones_genre_layout(tmp_path):
    sheets = {
        "CATÁLOGO COMPLETO": [
            ("TÍTULO", "AUTOR", "GÉNERO", "EDITORIAL", "PRECIOS", "STOCK", "Observaciones"),
            ("El astillero", "Onetti, Juan Carlos", "Novela", "Sudamericana", 29500, 3, "Cande"),
        ]
    }
    path = _xlsx_bytes(sheets, tmp_path / "catalog.xlsx")
    parsed = parse_workbook(path)
    row = parsed.sheets[0].rows[0]
    assert row.observaciones == "Cande"
    assert row.genre == "Novela"
    assert row.error is None


def test_parser_oportunidades_observaciones_none(tmp_path):
    sheets = {
        "Oportunidades": [
            ("Ejercicios de ocupación", "AA.VV.", "Ensayo", "VenteVeo", 1, 15000),
        ]
    }
    path = _xlsx_bytes(sheets, tmp_path / "catalog.xlsx")
    parsed = parse_workbook(path)
    row = parsed.sheets[0].rows[0]
    assert row.observaciones is None


# --------------------------------------------------------------------------
# Import applies observaciones (verbatim, empty -> Juli, no-column -> Juli)
# --------------------------------------------------------------------------

async def test_import_applies_observaciones_verbatim(auth_headers, session, client):
    cid = await _category_id(session)
    sheets = {
        "NOVELAS": [
            ("TÍTULO", "AUTOR", "EDITORIAL", "PRECIOS", "STOCK", "OBSERVACIONES"),
            ("Rayuela", "Cortázar, Julio", "Sudamericana", 29500, 3, "Juli y Cande"),
        ]
    }
    preview = await _preview(client, auth_headers, sheets)
    response = await _apply(client, auth_headers, preview)
    assert response.status_code == 200

    book = (await session.execute(select(Book))).scalar_one()
    assert book.observaciones == "Juli y Cande"


async def test_import_empty_observaciones_defaults_juli(auth_headers, session, client):
    await _category_id(session)
    sheets = {
        "NOVELAS": [
            ("TÍTULO", "AUTOR", "EDITORIAL", "PRECIOS", "STOCK", "OBSERVACIONES"),
            ("Rayuela", "Cortázar, Julio", "Sudamericana", 29500, 3, None),
        ]
    }
    preview = await _preview(client, auth_headers, sheets)
    response = await _apply(client, auth_headers, preview)
    assert response.status_code == 200

    book = (await session.execute(select(Book))).scalar_one()
    assert book.observaciones == "Juli"


async def test_import_oportunidades_defaults_juli(auth_headers, session, client):
    await _category_id(session, "OPORTUNIDADES")
    sheets = {
        "Oportunidades": [
            ("Ejercicios de ocupación", "AA.VV.", "Ensayo", "VenteVeo", 1, 15000),
        ]
    }
    preview = await _preview(client, auth_headers, sheets)
    response = await _apply(client, auth_headers, preview)
    assert response.status_code == 200

    book = (await session.execute(select(Book))).scalar_one()
    assert book.observaciones == "Juli"


# --------------------------------------------------------------------------
# Seed script backfill logic
# --------------------------------------------------------------------------

def test_seed_read_catalog_observaciones(tmp_path):
    sheets = {
        "NOVELAS": [
            ("TÍTULO", "AUTOR", "EDITORIAL", "PRECIO", "STOCK", "OBSERVACIONES"),
            ("Rayuela", "Cortázar, Julio", "Sudamericana", 29500, 3, "Juli y Cande"),
            ("1984", "Orwell, George", "La Pollera", 24000, 1, None),
        ],
        "Oportunidades": [
            ("Ejercicios", "AA.VV.", "Ensayo", "VenteVeo", 1, 15000),
        ],
    }
    path = _xlsx_bytes(sheets, tmp_path / "catalog.xlsx")
    mapping = seed.read_catalog_observaciones(path)
    assert mapping[("rayuela", "cortázar, julio", "sudamericana")] == "Juli y Cande"
    assert mapping[("1984", "orwell, george", "la pollera")] == "Juli"
    assert mapping[("ejercicios", "aa.vv.", "venteveo")] == "Juli"


def test_seed_read_sales_observaciones_julio_quirk(tmp_path):
    sheets = {
        "VENTAS JULIO 26": [
            (None, None, None, None, None, None, "Juli", "Cande"),
            ("Tsundoku", "Raito Pym, Taito", "Godot", 28900, 1, "Cande", None, 28900),
            ("Una casa", "Obligado, Clara", "Eme", 22000, 0, None, 22000, None),
        ],
        "VENTAS AGOSTO 26": [
            ("TÍTULO", "AUTOR", "EDITORIAL", "PRECIO", "STOCK", "OBSERVACIONES", "Método", "Fecha", "Juli", "Cande"),
            ("Señor Gato", "Blexbolex", "Libros del Zorro Rojo", 31400, 1, "Juli y Cande", None, None, 15700, 15700),
        ],
    }
    path = _xlsx_bytes(sheets, tmp_path / "ventas.xlsx")
    mapping = seed.read_sales_observaciones(path)
    assert mapping[("tsundoku", "raito pym, taito", "godot")] == "Cande"
    assert mapping[("una casa", "obligado, clara", "eme")] == "Juli"
    assert mapping[("señor gato", "blexbolex", "libros del zorro rojo")] == "Juli y Cande"


def test_seed_read_sales_non_julio_sheet_without_header_defaults_juli(tmp_path):
    # A non-JULIO sheet with no /observ/i header must NOT assume column 5.
    sheets = {
        "OTRA HOJA": [
            (None, None, None, None, None, None, "Juli", "Cande"),
            ("Libro X", "Autor Y", "Editorial Z", 1000, 1, "IGNORADO", None, 1000),
        ],
    }
    path = _xlsx_bytes(sheets, tmp_path / "ventas.xlsx")
    mapping = seed.read_sales_observaciones(path)
    assert mapping[("libro x", "autor y", "editorial z")] == "Juli"


async def test_seed_backfill_books(session):
    book = await _seed_book(session, title="Rayuela")
    before = (await session.execute(select(Book).where(Book.id == book.id))).scalar_one()
    mapping = {("rayuela", "julio cortázar", "sudamericana"): "Juli y Cande"}
    updated, skipped = await seed.backfill_book_observaciones(session, mapping)
    assert updated == 1
    assert skipped == 0

    refreshed = (await session.execute(select(Book).where(Book.id == book.id))).scalar_one()
    assert refreshed.observaciones == "Juli y Cande"
    # Price/stock/is_active untouched.
    assert refreshed.price == before.price
    assert refreshed.stock == before.stock
    assert refreshed.is_active == before.is_active


async def test_seed_backfill_books_skips_unknown(session):
    await _seed_book(session, title="Rayuela")
    mapping = {("otro", "autor", "editorial"): "Juli"}
    updated, skipped = await seed.backfill_book_observaciones(session, mapping)
    assert updated == 0
    assert skipped == 1

    refreshed = (await session.execute(select(Book))).scalar_one()
    assert refreshed.observaciones is None


async def test_seed_backfill_sales_uses_sales_excel_value(session):
    book = await _seed_book(session, title="Rayuela")
    # Book already has a catalog value, but the sales Excel should win.
    book.observaciones = "Cande"
    await session.commit()
    sale = await _seed_sale(session, book)
    mapping = {("rayuela", "julio cortázar", "sudamericana"): "Juli y Cande"}
    updated, not_found = await seed.backfill_sale_observaciones(session, mapping)
    assert updated == 1
    assert not_found == 0

    refreshed = (await session.execute(select(Sale).where(Sale.id == sale.id))).scalar_one()
    assert refreshed.observaciones == "Juli y Cande"


async def test_seed_backfill_sales_falls_back_to_book_observaciones(session):
    book = await _seed_book(session, title="Rayuela")
    book.observaciones = "Cande"
    await session.commit()
    sale = await _seed_sale(session, book)
    # The book is NOT in the sales Excel map.
    mapping = {("otro", "autor", "editorial"): "Juli"}
    updated, not_found = await seed.backfill_sale_observaciones(session, mapping)
    assert updated == 1
    assert not_found == 0

    refreshed = (await session.execute(select(Sale).where(Sale.id == sale.id))).scalar_one()
    assert refreshed.observaciones == "Cande"


async def test_seed_backfill_sales_null_when_no_source(session):
    book = await _seed_book(session, title="Rayuela")
    sale = await _seed_sale(session, book)
    # Book.observaciones is NULL and the book is not in the sales Excel map.
    mapping = {("otro", "autor", "editorial"): "Juli"}
    updated, not_found = await seed.backfill_sale_observaciones(session, mapping)
    assert updated == 0
    assert not_found == 1

    refreshed = (await session.execute(select(Sale).where(Sale.id == sale.id))).scalar_one()
    assert refreshed.observaciones is None
