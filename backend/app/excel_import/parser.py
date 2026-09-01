"""Read-only openpyxl parser with per-sheet layout detection.

Two layouts (REQ-IMP-1):
- header sheets: row 1 = ``TÍTULO | AUTOR | EDITORIAL | PRECIO(S) | STOCK``
- OPORTUNIDADES: no header row. The real file mixes two sub-layouts within the
  same sheet, so the column map is chosen per row from the non-empty column
  count (source of truth: the catalog file itself):
  - 6 columns: ``TÍTULO | AUTOR | GÉNERO | EDITORIAL | STOCK | PRECIO``
  - 5 columns: ``TÍTULO | AUTOR | EDITORIAL | PRECIO | STOCK`` (no genre)

Rows are normalized (strip whitespace, keep display case), prices become
``Decimal`` (plain ARS ints or ``$X.XXX,XX`` formatted strings), stock becomes
``int``, and malformed rows carry a structured error instead of aborting the
whole file.
"""

from __future__ import annotations

import itertools
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .normalizer import (
    DEFAULT_CATEGORIES,
    GENRE_FALLBACK_CATEGORY,
    SKIP_SHEETS,
    category_for_genre,
    category_for_sheet,
    has_genre_column,
    is_header_row,
    normalize_sheet_name,
    observaciones_column_index,
)

HEADER_COLUMNS = ["title", "author", "editorial", "price", "stock"]
# Genre-layout header sheets carry a GÉNERO column at position 3 and EDITORIAL
# at position 4 (CATÁLOGO COMPLETO): ``TÍTULO | AUTOR | GÉNERO | EDITORIAL |
# PRECIO(S) | STOCK``.
GENRE_LAYOUT_COLUMNS = ["title", "author", "genre", "editorial", "price", "stock"]
OPORTUNIDADES_COLUMNS = ["title", "author", "genre", "editorial", "stock", "price"]
# Only the leading columns are meaningful; trailing empty columns (this file
# reports max_column 24 on some sheets) must not be parsed as data.
MAX_COLUMNS = 10

XLSX_EXTENSIONS = (".xlsx", ".xlsm")


class ExcelImportError(Exception):
    """Base error for malformed/unsupported Excel input (mapped to HTTP 400)."""


class UnsupportedFileError(ExcelImportError):
    """File is not a readable .xlsx workbook."""


class EmptyWorkbookError(ExcelImportError):
    """Workbook exists but exposes no worksheets."""


@dataclass
class ParsedRow:
    """One data row.

    ``error`` is set (instead of ``price``/``stock``) when malformed. ``skip_reason``
    marks data-entry noise that is silently skipped (counted as skips, never errors)
    and never reaches apply.
    """

    sheet: str
    category: str | None
    title: str
    author: str
    editorial: str
    genre: str | None
    price: Decimal | None
    stock: int | None
    observaciones: str | None
    row_number: int
    error: str | None = None
    skip_reason: str | None = None


@dataclass
class ParsedSheet:
    """Parsed worksheet: detected layout + normalized rows."""

    name: str
    category: str | None
    has_header: bool
    genre_driven: bool = False
    rows: list[ParsedRow] = field(default_factory=list)


@dataclass
class ParsedWorkbook:
    """Parsed workbook: filename + one :class:`ParsedSheet` per worksheet."""

    filename: str
    sheets: list[ParsedSheet] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return any(row.error is not None for sheet in self.sheets for row in sheet.rows)


def _clean(value) -> str | None:
    """Return a stripped string cell value, or ``None`` when empty."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_datetime(value) -> bool:
    return isinstance(value, datetime)


def _normalize_ars_price(text: str) -> str:
    """Turn an ARS price cell into a ``Decimal``-parseable string.

    Accepts plain numbers (``29500``), ``$``-prefixed amounts, and formatted
    values such as ``$24.000,00`` (dot = thousands, comma = decimals).
    """
    text = text.replace("$", "").strip()
    if "," in text:
        return text.replace(".", "").replace(",", ".")
    if "." in text:
        integer, _, fractional = text.partition(".")
        if len(fractional) == 3 and "." not in fractional:
            return integer + fractional
    return text


def _parse_price(value) -> Decimal | None:
    """Parse a price cell into a non-negative ``Decimal`` (2 decimals)."""
    if isinstance(value, bool):
        return None
    try:
        price = Decimal(_normalize_ars_price(str(value).strip()))
    except (InvalidOperation, ValueError):
        return None
    if not price.is_finite() or price < 0:
        return None
    return price.quantize(Decimal("0.01"))


def _parse_stock(value) -> int | None:
    """Parse a stock cell into a non-negative ``int`` (tolerates ``5.0``)."""
    if isinstance(value, bool):
        return None
    try:
        stock = int(str(value).strip())
    except ValueError:
        try:
            stock = float(value)
        except (ValueError, TypeError):
            return None
        if not stock.is_integer():
            return None
        stock = int(stock)
    return stock if stock >= 0 else None


def _columns_for_no_header_row(cells: list[str | None]) -> list[str]:
    """Pick the OPORTUNIDADES column map from the non-empty column count."""
    non_empty = sum(1 for cell in cells if cell)
    if non_empty == 5:
        return HEADER_COLUMNS
    return OPORTUNIDADES_COLUMNS


def _build_row(
    sheet_name: str,
    category: str | None,
    columns: list[str],
    cells: list,
    row_number: int,
    *,
    genre_driven: bool = False,
    available_categories: Iterable[str] = (),
    obs_index: int | None = None,
) -> ParsedRow:
    values = dict(zip(columns, cells))
    title_raw = values.get("title")
    author_raw = values.get("author")
    editorial_raw = values.get("editorial")
    genre_raw = values.get("genre")
    price_raw = values.get("price")
    stock_raw = values.get("stock")
    observaciones_raw = cells[obs_index] if obs_index is not None and obs_index < len(cells) else None

    # Clean and parse every field first: skip and error rows both carry the
    # parsed values so summary accounting stays accurate.
    title = _clean(title_raw)
    author = _clean(author_raw)
    editorial = _clean(editorial_raw)
    genre = _clean(genre_raw)
    observaciones = _clean(observaciones_raw)
    price = _parse_price(price_raw) if price_raw is not None else None
    stock = _parse_stock(stock_raw) if stock_raw is not None else None

    if genre_driven:
        resolved = category_for_genre(genre, available_categories)
        row_category = resolved if resolved is not None else GENRE_FALLBACK_CATEGORY
    else:
        row_category = category

    # Data-entry noise heuristics: silently skip (counted as skips, never as
    # errors) BEFORE genuine validation-error accumulation.
    if _is_datetime(title_raw):
        return ParsedRow(
            sheet=sheet_name,
            category=row_category,
            title=title or "",
            author=author or "",
            editorial=editorial or "",
            genre=genre,
            price=price,
            stock=stock,
            observaciones=observaciones,
            row_number=row_number,
            skip_reason="Unexpected date value in title",
        )
    if (
        not title
        and not author
        and not editorial
        and isinstance(price_raw, str)
        and price_raw.strip()
        and price is None
    ):
        return ParsedRow(
            sheet=sheet_name,
            category=row_category,
            title=title or "",
            author=author or "",
            editorial=editorial or "",
            genre=genre,
            price=price,
            stock=stock,
            observaciones=observaciones,
            row_number=row_number,
            skip_reason="footer/summary row",
        )

    errors: list[str] = []
    if _is_datetime(author_raw):
        errors.append("Unexpected date value in author")
    if _is_datetime(editorial_raw):
        errors.append("Unexpected date value in editorial")
    if _is_datetime(genre_raw):
        errors.append("Unexpected date value in genre")

    if not title:
        errors.append("Missing title")
    if not author:
        errors.append("Missing author")
    if not editorial:
        errors.append("Missing editorial")

    if price_raw is None:
        errors.append("Missing price")
    elif price is None:
        errors.append(f"Invalid price {price_raw!r}")

    if stock_raw is None:
        errors.append("Missing stock")
    elif stock is None:
        errors.append(f"Invalid stock {stock_raw!r}")

    if row_category is None:
        errors.append(f"No category mapped for sheet {sheet_name!r}")

    return ParsedRow(
        sheet=sheet_name,
        category=row_category,
        title=title or "",
        author=author or "",
        editorial=editorial or "",
        genre=genre,
        price=price,
        stock=stock,
        observaciones=observaciones,
        row_number=row_number,
        error="; ".join(errors) if errors else None,
    )


def _parse_sheet(
    worksheet, category: str | None, available_categories: Iterable[str]
) -> ParsedSheet:
    iterator = worksheet.iter_rows(values_only=True)
    first = next(iterator, None)
    if first is None:
        return ParsedSheet(name=worksheet.title, category=category, has_header=False)

    first_cleaned = [_clean(cell) for cell in first[:MAX_COLUMNS]]
    has_header = is_header_row(first_cleaned)
    genre_driven = has_header and has_genre_column(first_cleaned)
    obs_index = observaciones_column_index(first_cleaned) if has_header else None
    data_iterator = iterator if has_header else itertools.chain([first], iterator)

    rows: list[ParsedRow] = []
    row_number = 2 if has_header else 1
    for values in data_iterator:
        cells = [_clean(cell) for cell in values[:MAX_COLUMNS]]
        if all(cell is None for cell in cells):
            row_number += 1
            continue
        if genre_driven:
            columns = GENRE_LAYOUT_COLUMNS
        elif has_header:
            columns = HEADER_COLUMNS
        else:
            columns = _columns_for_no_header_row(cells)
        rows.append(
            _build_row(
                worksheet.title,
                category,
                columns,
                values[:MAX_COLUMNS],
                row_number,
                genre_driven=genre_driven,
                available_categories=available_categories,
                obs_index=obs_index,
            )
        )
        row_number += 1

    return ParsedSheet(
        name=worksheet.title,
        category=category,
        has_header=has_header,
        genre_driven=genre_driven,
        rows=rows,
    )


def parse_workbook(
    source: bytes | Path | str,
    *,
    filename: str | None = None,
    available_categories: list[str] | None = None,
) -> ParsedWorkbook:
    """Parse an ``.xlsx`` workbook from bytes or a path.

    Raises :class:`ExcelImportError` subclasses for unsupported/empty files.
    """
    if isinstance(source, (str, Path)):
        path = Path(source)
        filename = filename or path.name
        stream: BytesIO | None = None
    else:
        filename = filename or "catalog.xlsx"
        stream = BytesIO(source)

    if filename and not filename.lower().endswith(XLSX_EXTENSIONS):
        raise UnsupportedFileError(
            f"Unsupported file type: {filename!r}; expected an .xlsx file"
        )

    try:
        workbook = load_workbook(
            stream if stream is not None else str(path),
            read_only=True,
            data_only=True,
        )
    except (InvalidFileException, zipfile.BadZipFile, OSError, KeyError, ValueError) as exc:
        raise UnsupportedFileError(
            "Could not read the file as an Excel workbook"
        ) from exc

    if not workbook.worksheets:
        raise EmptyWorkbookError("Workbook contains no worksheets")

    available = available_categories or list(DEFAULT_CATEGORIES)
    sheets = [
        _parse_sheet(worksheet, category_for_sheet(worksheet.title, available), available)
        for worksheet in workbook.worksheets
        if normalize_sheet_name(worksheet.title) not in SKIP_SHEETS
    ]
    return ParsedWorkbook(filename=filename, sheets=sheets)
